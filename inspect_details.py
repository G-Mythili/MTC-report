from openpyxl import load_workbook
wb = load_workbook(r"e:\2026 Report Famat\Final correct.xlsx")
ws = wb.active

print("Checking Rows 4, 5, 6:")
for r in [4, 5, 6]:
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        print(f"Row {r}, Col {c} ({cell.coordinate}): '{val}'")

print("\nChecking Rows 27 to 40:")
for r in range(27, 41):
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val:
            print(f"Row {r}, Col {c} ({cell.coordinate}): '{val}'")

print("\nChecking Rows 41 to 43:")
for r in range(41, 44):
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val:
            print(f"Row {r}, Col {c} ({cell.coordinate}): '{val}'")

