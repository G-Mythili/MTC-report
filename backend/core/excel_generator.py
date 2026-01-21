from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
import os
import io

class ExcelGenerator:
    def __init__(self, settings):
        self.settings = settings
        self.b_style = settings.get("border_style", "thin")
        self.font_family = settings.get("font_family", "Calibri")
        self.font_size = settings.get("font_size", 10)
        self.header_color = settings.get("header_fill_color", "#d9e1f2").replace("#", "")
        
        # Styles
        self.side_obj = Side(border_style=self.b_style, color="000000") if self.b_style != "none" else Side()
        self.full_border = Border(top=self.side_obj, left=self.side_obj, right=self.side_obj, bottom=self.side_obj)
        self.custom_font = Font(name=self.font_family, size=self.font_size)
        self.header_fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

        # User Specific Row Mappings (Fixed to match M537 template)
        self.row_map = {
            "Carbon": 14,
            "Silicon": 15,
            "Manganese": 16,
            "Phosphorus": 17,
            "Sulphur": 18,
            "Copper": 19,
            "Nickel": 20,
            "Chromium": 21,
            "Moly": 22,
            "Magnesium": 23,
            "CE": 24,
            "Tin": 26, # User requested Row 26 for Tin
        }
        self.BASE = 2 # Col B (Spec)

    def generate(self, template_path, data):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found: {template_path}")

        wb = load_workbook(template_path)
        ws = wb.active
        
        # 1. Surgical Sanitization (Only Column E onwards)
        # We NO LONGER unmerge everything. We preserve A-D layout.
        for sheet in wb.worksheets:
            # Set specific default widths for A-D
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 38
            sheet.column_dimensions['C'].width = 17
            sheet.column_dimensions['D'].width = 24
            
            # Wipe Columns E through Z surgically
            for r in range(1, 251):
                # Targeted unmerge ONLY if a merge range touches Column E or beyond
                for rng in list(sheet.merged_cells.ranges):
                    if rng.max_col >= 5:
                        try: sheet.unmerge_cells(str(rng))
                        except: pass
                
                for c in range(5, 51): # E to AZ
                    cell = sheet.cell(row=r, column=c)
                    cell.value = None
                    cell.border = Border() 
                    cell.fill = PatternFill(fill_type=None)
            
            # Step 1.1: Surgical Unmerge for Data Rows (14-26)
            for rng in list(sheet.merged_cells.ranges):
                if rng.min_row >= 14 and rng.max_row <= 26 and rng.min_col <= 4:
                    try: sheet.unmerge_cells(str(rng))
                    except: pass

        # 2. Update Header Values (Concatenate labels to preserve template text)
        invoice_val = data.get("invoice_no", "")
        qty_val = data.get("qty", "") # Removed replace("no's", "Nos") as user wants 'no's'
        date_val = data.get("date", "")

        self._safe_write(ws, 4, 3, f"Invoice No : {invoice_val}")
        self._safe_write(ws, 5, 3, f"Despatch Quantity:  {qty_val}")
        self._safe_write(ws, 6, 3, f"Dispatch Date : {date_val}")
        self._safe_write(ws, 8, 2, data.get("part_details", "")) 

        # Clear Dynamic Data Rows COMPLETELY (Rows 14-26 and 28-31, Col 1-4)
        for row_to_clear in list(range(14, 27)) + list(range(28, 32)):
            for c_idx in range(1, 5):
                self._safe_write(ws, row_to_clear, c_idx, None)
        
        # Heat Nos (Row 13)
        self._write_styled(ws, 13, self.BASE+1, data.get("heat1", ""), fill=True)
        self._write_styled(ws, 13, self.BASE+2, data.get("heat2", ""), fill=True)

        # 3. Logo (Insert if missing)
        self._insert_logo(ws)

        # 4. Chemistry (Dynamic Write & Row Hiding)
        chem_data = data.get("chemistry", [])
        written_rows = set()
        for item in chem_data:
            elem_name = item.get("Element", "")
            target_row = None
            for key, row_idx in self.row_map.items():
                if key.lower() in elem_name.lower():
                    target_row = row_idx
                    break
            
            if target_row:
                written_rows.add(target_row)
                # Write Label (Col 1), Spec (Col 2), and Observations (Col 3, 4)
                # V15: Carbon row (Row 14) should NOT have the heat value fill (force white)
                if target_row == 14:
                    self._write_styled(ws, target_row, 1, elem_name, align='left', fill=True)
                    self._write_styled(ws, target_row, 2, item.get("Spec", ""))
                    self._write_styled(ws, target_row, 3, self._fmt(item.get('heat1_val', '')), fill=True)
                    self._write_styled(ws, target_row, 4, self._fmt(item.get('heat2_val', '')), fill=True)
                else:
                    self._write_styled(ws, target_row, 1, elem_name, align='left')
                    self._write_styled(ws, target_row, 2, item.get("Spec", ""))
                    self._write_styled(ws, target_row, 3, self._fmt(item.get('heat1_val', '')), fill=True)
                    self._write_styled(ws, target_row, 4, self._fmt(item.get('heat2_val', '')), fill=True)

        # Hide empty rows in chemistry block (14-26) to stop "empty row" issue
        for r in range(14, 27):
            if r not in written_rows:
                ws.row_dimensions[r].hidden = True
            else:
                ws.row_dimensions[r].hidden = False

        # 5. Mechanical (Dynamic Write - Index Based for Robustness)
        mech_data = data.get("mechanical", [])
        # Map the first 4 items to rows 28, 29, 30, 31 respectively
        for i, row in enumerate(mech_data[:4]):
            target_row = 28 + i
            param = row.get("Parameter", "")
            
            # Write Label (Col 1), Spec (Col 2), and observations (Col 3)
            self._write_styled(ws, target_row, 1, param, align='left')
            self._write_styled(ws, target_row, 2, row.get("Spec", ""))
            
            # V15: Merge C and D for observations in Rows 28-31
            try:
                # Always unmerge first to avoid conflicts if template has partial merges
                ws.unmerge_cells(start_row=target_row, start_column=3, end_row=target_row, end_column=4)
            except: pass
            try:
                ws.merge_cells(start_row=target_row, start_column=3, end_row=target_row, end_column=4)
            except: pass
            
            self._write_styled(ws, target_row, 3, row.get('heat1_val', ''), fill=True)

        # --- Footer & Grade ---
        grade = data.get("grade", "GRADE 4512")
        conc_text = f"Conclusion: The above material is satisfactory to Ductile iron J434C GRADE {grade}."
        conc_row = 42 
        self._write_styled(ws, conc_row, 1, conc_text, align='left')
        
        # --- Final Print Configuration ---
        try:
            # Orientation and Paper Size
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            
            # Explicit Print Area (Columns A-D)
            ws.print_area = 'A1:D50'
            
            # Robust scaling setup
            # We must force the worksheet toFit to Page
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0 # Automatic 
            
            # Center horizontally and reduce margins for maximum content size
            ws.print_options.horizontalCentered = True
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            
            # V15: Apply Thick Outer Border (A1:D50)
            thick_side = Side(style='thick')
            # Top
            for c in range(1, 5):
                cell = ws.cell(row=1, column=c)
                cell.border = Border(top=thick_side, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
            # Bottom
            for c in range(1, 5):
                cell = ws.cell(row=50, column=c)
                cell.border = Border(bottom=thick_side, left=cell.border.left, right=cell.border.right, top=cell.border.top)
            # Left
            for r in range(1, 51):
                cell = ws.cell(row=r, column=1)
                cell.border = Border(left=thick_side, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
            # Right
            for r in range(1, 51):
                cell = ws.cell(row=r, column=4)
                cell.border = Border(right=thick_side, left=cell.border.left, top=cell.border.top, bottom=cell.border.bottom)

        except Exception as e:
            # Fallback for older openpyxl or missing parents
            try:
                ws.page_setup.fitToWidth = 1
                ws.print_area = 'A1:D50'
            except: pass

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _fmt(self, val):
        if val is None or val == "": return ""
        try:
            return f"{float(val):.2f}%"
        except:
            return str(val)

    def _get_target_cell(self, ws, row, col):
        """Helper to find the top-left cell of a merge range."""
        cell = ws.cell(row=row, column=col)
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                return ws.cell(row=rng.min_row, column=rng.min_col)
        return cell

    def _write_styled(self, ws, row, col, value, align='center', fill=False, is_header=False):
        if col > 4: return # STRICT ABCD BOUNDARY
        try:
            target = self._get_target_cell(ws, row, col)
            target.value = value
            target.border = self.full_border
            target.font = self.custom_font
            if align == 'center': target.alignment = self.center_align
            elif align == 'left': target.alignment = self.left_align
            
            if is_header: target.fill = self.header_fill
            elif fill: target.fill = self.white_fill
            else: target.fill = PatternFill(fill_type=None) # Explicitly remove any fill (V15)
        except: pass

    def _safe_write(self, ws, row, col, value):
        if col > 4: return # STRICT ABCD BOUNDARY
        try:
            target = self._get_target_cell(ws, row, col)
            target.value = value
        except: pass

    def _write_footer(self, ws, row):
        # We trust the template for most styles, just update names
        self._safe_write(ws, row+1, 2, "G.Mythili")
        self._safe_write(ws, row+1, 4, "T.Thirugnanam")

    def _insert_logo(self, ws):
        try:
            # Only insert if no images exist in the top-left area
            if not any(img.anchor == 'A1' for img in ws._images):
                logo_path = os.path.join(os.getcwd(), "logo.png")
                if os.path.exists(logo_path):
                    img = Image(logo_path)
                    img.height, img.width = 120, 75
                    ws.add_image(img, 'A1')
        except: pass

