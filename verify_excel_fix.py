from backend.core.excel_generator import ExcelGenerator
import os
import json

settings = {
    "border_style": "thin",
    "font_family": "Calibri",
    "font_size": 10,
    "header_fill_color": "#d9e1f2"
}

data = {
    "invoice_no": "INV-123",
    "qty": "100 Nos",
    "date": "2026-01-12",
    "part_details": "Test Part Details",
    "heat1": "HEAT-001",
    "heat2": "HEAT-002",
    "chemistry": [
        {"Element": "Carbon", "heat1_val": "3.5", "heat2_val": "3.4"},
        {"Element": "Silicon", "heat1_val": "2.1", "heat2_val": "2.2"},
        {"Element": "Tin", "heat1_val": "0.1", "heat2_val": "0.1"}
    ],
    "mechanical": [
        {"Parameter": "Hardness", "heat1_val": "200 BHN"},
        {"Parameter": "Tensile Strength", "heat1_val": "500 Mpa"},
        {"Parameter": "Yield Strength", "heat1_val": "300 Mpa"},
        {"Parameter": "Elongation", "heat1_val": "15%"}
    ],
    "grade": "4512"
}

template_path = r"e:\2026 Report Famat\Final correct.xlsx"
generator = ExcelGenerator(settings)
output_bytes = generator.generate(template_path, data)

output_path = r"e:\2026 Report Famat\verification_output.xlsx"
with open(output_path, "wb") as f:
    f.write(output_bytes)

# Verify
from openpyxl import load_workbook
wb = load_workbook(output_path)
ws = wb.active

errors = []

# 1. Check Merges
merges = [str(r) for r in ws.merged_cells.ranges]
expected_merges = ["A2:B2", "C4:D4"] # Removed A14:D14 as we now unmerge it for data
for em in expected_merges:
    if not any(em in m for m in merges):
        errors.append(f"Missing expected merge range: {em}")

# 1.1 Check Print Setup
if 'A$1:$D$50' not in str(ws.print_area):
    errors.append(f"Print Area mismatch: {ws.print_area}, expected A1:D50")

if not ws.sheet_properties.pageSetUpPr.fitToPage:
    errors.append("FitToPage not set in sheet_properties")

# 2. Check Column E
for r in range(1, 100):
    val = ws.cell(row=r, column=5).value
    if val is not None:
        errors.append(f"Column E is NOT empty at Row {r}: {val}")

# 3. Check Concatenated Headers
h4 = ws.cell(row=4, column=3).value
if "Invoice No : INV-123" not in str(h4):
    errors.append(f"Row 4 Header mismatch: {h4}")

h5 = ws.cell(row=5, column=3).value
if "Despatch Quantity:  100 Nos" not in str(h5):
    errors.append(f"Row 5 Header mismatch: {h5}")

# 4. Check Chemistry Row Mapping & Clearing
c14 = ws.cell(row=14, column=3).value
if "3.50%" not in str(c14):
    errors.append(f"Carbon not found at Row 14: {c14}")

c26 = ws.cell(row=26, column=3).value
if "0.10%" not in str(c26):
    errors.append(f"Tin not found at Row 26: {c26}")

c21 = ws.cell(row=21, column=3).value
if c21 is not None and not ws.row_dimensions[21].hidden:
    errors.append(f"Slot Row 21 (Nickel) should be HIDDEN if empty, but found: {c21}")

# 5. Check Mechanical Row Mapping & Merges
m28 = ws.cell(row=28, column=3).value
if "200 BHN" not in str(m28):
    errors.append(f"Hardness not found at Row 28: {m28}")

# Check if C28:D28 is merged
mechanical_merges = ["C28:D28", "C29:D29", "C30:D30", "C31:D31"]
for mm in mechanical_merges:
    if not any(mm in m for m in merges):
         errors.append(f"Missing expected mechanical merge: {mm}")

# 6. Check Aesthetic Refinements (V15)
# Carbon row background (should be None or white)
carbon_fill = ws.cell(row=14, column=1).fill.start_color.index
if carbon_fill not in ['00000000', 'FFFFFFFF', '00FFFFFF']:
    errors.append(f"Carbon row (14) background should be white/None, but found: {carbon_fill}")

# Thick border check (A1)
border_left = ws.cell(row=1, column=1).border.left.style
if border_left != 'thick':
    errors.append(f"Outer border at A1 (left) should be thick, but found: {border_left}")

if not errors:
    print("VERIFICATION SUCCESSFUL: Aesthetics V15 (White Carbon, Merged Mechanical, Thick Border)")
else:
    print("VERIFICATION FAILED:")
    for e in errors:
        print(f" - {e}")

