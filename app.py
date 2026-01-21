import streamlit as st
import pandas as pd
import os
import io
import json
import webbrowser
from openpyxl import load_workbook
import streamlit.components.v1 as components
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode

st.set_page_config(page_title="Spectro Report Viewer", layout="wide")

# --- Settings Persistence ---
SETTINGS_FILE = "settings.json"
DEFAULT_MTC_TEMPLATE = r"E:\Final correct.xlsx"

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r") as f:
                return json.load(f)
        except:
            pass
    return {
        "chem_title": "1. Chemical composition",
        "mech_title": "2. Mechanical Properties",
        "micro_title": "3. Microstructure",
        "matrix_title": "3.1 Matrix",
        "header_align": "center",
        "mtc_template_path": DEFAULT_MTC_TEMPLATE,
        "header_fill_color": "#d9e1f2",
        "border_style": "thin",
        "font_family": "Calibri",
        "font_size": 10
    }

def save_settings(settings):
    try:
        with open(SETTINGS_FILE, "w") as f:
            json.dump(settings, f, indent=4)
    except Exception as e:
        st.error(f"Error saving settings: {e}")

# Initialize session state for settings
if "app_settings" not in st.session_state:
    st.session_state.app_settings = load_settings()

def render_report_viewer():
    st.header("Daily Report")
    
    # Default file path
    DEFAULT_FILE_PATH = r"E:\SPECTRO-09-01-2026.xlsx"
    
    # File Uploader
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"], key="report_uploader")
    
    data = None
    source_name = ""
    
    if uploaded_file is not None:
        try:
            data = pd.read_excel(uploaded_file)
            source_name = uploaded_file.name
            st.success(f"Loaded uploaded file: {source_name}")
        except Exception as e:
            st.error(f"Error reading uploaded file: {e}")
    else:
        # Try loading the default file
        if os.path.exists(DEFAULT_FILE_PATH):
            try:
                data = pd.read_excel(DEFAULT_FILE_PATH)
                source_name = DEFAULT_FILE_PATH
            except Exception as e:
                st.error(f"Error reading default file at {DEFAULT_FILE_PATH}: {e}")
        else:
            st.warning(f"Default file not found at: {DEFAULT_FILE_PATH}")
            st.markdown("Please upload a file to view.")
    
    
    # Display Data with AgGrid
    if data is not None:
        st.subheader(f"Report Data: {source_name}")
    
        # Add S.No column
        if 'S.No' not in data.columns:
            data.insert(0, 'S.No', range(1, 1 + len(data)))
    
        # Calculate CE% = C + (Si / 3) + P
        # We look for exact matches of standard chemistry headers
        c_col = "C [%]"
        si_col = "Si [%]"
        p_col = "P [%]"
        
        if all(col in data.columns for col in [c_col, si_col, p_col]):
            try:
                # Ensure columns are numeric
                c_val = pd.to_numeric(data[c_col], errors='coerce')
                si_val = pd.to_numeric(data[si_col], errors='coerce')
                p_val = pd.to_numeric(data[p_col], errors='coerce')
                
                # Calculate formula
                ce_val = c_val + (si_val / 3) + p_val
                
                # Round C, Si, and CE to 2 decimal places as requested
                data[c_col] = c_val.round(2)
                data[si_col] = si_val.round(2)
                ce_val = ce_val.round(2)
                
                # Insert the CE% column (e.g., after standard chemistry if possible, or just append)
                # User requested to move CE% next to Si% (after Si)
                if si_col in data.columns:
                    loc_index = data.columns.get_loc(si_col) + 1
                elif c_col in data.columns:
                    loc_index = data.columns.get_loc(c_col) + 1
                else:
                    loc_index = 1
                    
                if "CE%" not in data.columns:
                    data.insert(loc_index, "CE%", ce_val)
                else:
                    data["CE%"] = ce_val # Update if exists
                
            except Exception as e:
                st.error(f"Error calculating CE%: {e}")
    
        # Sidebar: Column Selection
        st.sidebar.markdown("---")
        st.sidebar.header("Report Settings")
        
        # --- Formatting Settings Expander ---
        with st.sidebar.expander("🛠️ Formatting Settings", expanded=False):
            st.subheader("Section Titles")
            s = st.session_state.app_settings
            
            new_chem = st.text_input("Chemistry Title", value=s["chem_title"])
            new_mech = st.text_input("Mechanical Title", value=s["mech_title"])
            new_micro = st.text_input("Microstructure Title", value=s["micro_title"])
            new_matrix = st.text_input("Matrix Title", value=s["matrix_title"])
            
            st.subheader("Header Alignment")
            new_align = st.selectbox("Text Alignment", options=["left", "center", "right"], index=["left", "center", "right"].index(s.get("header_align", "center")))
            
            st.subheader("Visual Styling")
            new_color = st.color_picker("Header Background Color", value=s.get("header_fill_color", "#d9e1f2"))
            new_border = st.selectbox("Border Style", options=["thin", "medium", "thick", "none"], index=["thin", "medium", "thick", "none"].index(s.get("border_style", "thin")))
            new_font = st.selectbox("Font Family", options=["Calibri", "Arial", "Times New Roman"], index=["Calibri", "Arial", "Times New Roman"].index(s.get("font_family", "Calibri")))
            new_size = st.slider("Font Size", 8, 14, value=s.get("font_size", 10))

            st.subheader("Template Management")
            current_tpl_path = st.text_input("MTC Template Path", value=s.get("mtc_template_path", DEFAULT_MTC_TEMPLATE))
            uploaded_template = st.file_uploader("Upload New MTC Template", type=["xlsx"], key="template_upload")
            
            if st.button("Save Settings as Default"):
                s["chem_title"] = new_chem
                s["mech_title"] = new_mech
                s["micro_title"] = new_micro
                s["matrix_title"] = new_matrix
                s["header_align"] = new_align
                s["header_fill_color"] = new_color
                s["border_style"] = new_border
                s["font_family"] = new_font
                s["font_size"] = new_size
                s["mtc_template_path"] = current_tpl_path
                
                if uploaded_template:
                    # Save the template locally
                    temp_path = os.path.join(os.getcwd(), uploaded_template.name)
                    with open(temp_path, "wb") as f:
                        f.write(uploaded_template.getbuffer())
                    s["mtc_template_path"] = temp_path
                    st.success(f"Template saved: {uploaded_template.name}")
                
                save_settings(s)
                st.session_state.app_settings = s
                st.success("Settings saved successfully!")
                st.rerun()

        # Deduplication Option
        deduplicate = st.sidebar.checkbox("Clean & Deduplicate (Max 2)", value=False, key="deduplicate")
        
        if deduplicate:
            # Identify columns
            sample_col = next((col for col in data.columns if "sample" in col.lower() and "id" in col.lower()), None)
            heat_col = next((col for col in data.columns if "heat" in col.lower() and "no" in col.lower()), None)
            
            # Clean Sample Id if found
            if sample_col:
                data[sample_col] = data[sample_col].astype(str).str.replace(r'-\d+$', '', regex=True)
            
            # Clean Heat No if found
            if heat_col:
                data[heat_col] = data[heat_col].astype(str).str.replace(r'-\d+$', '', regex=True)

            # Choose Deduplication Key
            options = []
            if sample_col: options.append("Sample Id")
            if heat_col: options.append("Heat No")
            
            target_col_name = None
            if options:
                dedup_choice = st.sidebar.radio("Group By (Max 2):", options, index=0)
                if dedup_choice == "Sample Id": target_col_name = sample_col
                if dedup_choice == "Heat No": target_col_name = heat_col
                
                if target_col_name:
                    # Apply Top 2 Logic
                    data = data.groupby(target_col_name).head(2).reset_index(drop=True)
                    
                    # Re-generate S.No
                    if 'S.No' in data.columns:
                        data['S.No'] = range(1, 1 + len(data))
            else:
                st.sidebar.warning("Could not find 'Sample Id' or 'Heat No' to deduplicate.")

        # --- Saved Formats Logic ---
        st.sidebar.markdown("---")
        st.sidebar.subheader("Saved Formats")
        
        FORMATS_FILE = "report_formats.json"
        
        def load_formats():
            if os.path.exists(FORMATS_FILE):
                try:
                    with open(FORMATS_FILE, "r") as f:
                        return json.load(f)
                except: return {}
            return {}
            
        def save_formats_to_file(formats):
             with open(FORMATS_FILE, "w") as f:
                 json.dump(formats, f)

        saved_formats = load_formats()
        
        # Save Current
        with st.sidebar.expander("Save Current View"):
            new_format_name = st.text_input("Format Name", placeholder="e.g. Morning Report")
            if st.button("Save Format"):
                if new_format_name:
                    # Save currently hidden columns
                    # We utilize the session state which tracks the current hidden cols
                    current_hidden = st.session_state.get("hidden_columns", [])
                    saved_formats[new_format_name] = current_hidden
                    save_formats_to_file(saved_formats)
                    st.success(f"Saved: {new_format_name}")
                    st.rerun()
                else:
                    st.error("Enter a name.")
                    
        # Load Format
        if saved_formats:
            selected_format = st.sidebar.selectbox("Select Format to Load", options=["Select..."] + list(saved_formats.keys()))
            
            if selected_format != "Select...":
                 col_load, col_del = st.sidebar.columns([2, 1])
                 with col_load:
                     if st.button("Load View"):
                         st.session_state.hidden_columns = saved_formats[selected_format]
                         st.session_state.select_all_cols = False # Usually we are subsetting
                         st.toast(f"Loaded: {selected_format}")
                         st.rerun()
                 with col_del:
                     if st.button("Del"):
                         del saved_formats[selected_format]
                         save_formats_to_file(saved_formats)
                         st.rerun()
    
        # Create a dataframe for column toggles
        all_columns = data.columns.tolist()
        
        # --- Persistence Logic ---
        if "hidden_columns" not in st.session_state:
            st.session_state.hidden_columns = []
            
        def toggle_all():
            if st.session_state.select_all_cols:
                st.session_state.hidden_columns = []
            else:
                st.session_state.hidden_columns = all_columns[:] # Hide all

        # Toggle to select/deselect all
        # We assume True by default if not set
        if "select_all_cols" not in st.session_state:
            st.session_state.select_all_cols = True

        select_all = st.sidebar.checkbox(
            "Select All Columns", 
            value=st.session_state.select_all_cols, 
            key="select_all_cols", 
            on_change=toggle_all
        )
        
        # Calculate visibility based on hidden_columns state
        # New columns in new files defaut to Visible (True)
        visibility_status = [col not in st.session_state.hidden_columns for col in all_columns]
        
        col_df = pd.DataFrame({
            "Column Name": all_columns,
            "Visible": visibility_status
        })
    
        # Display as a table with checkboxes
        st.sidebar.markdown("Toggle columns visibility:")
        
        with st.sidebar.form("col_vis_form"):
            edited_col_df = st.data_editor(
                col_df,
                column_config={
                    "Column Name": st.column_config.TextColumn("Column", disabled=True),
                    "Visible": st.column_config.CheckboxColumn("Show?", default=True)
                },
                hide_index=True,
                use_container_width=True,
                height=600, 
                key="col_editor"
            )
            submit_cols = st.form_submit_button("Apply Visibility Changes")
        
        # Update persistent state based on editor changes (Only on Submit)
        if submit_cols:
            current_hidden = edited_col_df[~edited_col_df["Visible"]]["Column Name"].tolist()
            st.session_state.hidden_columns = current_hidden
        
        # Filter data based on session state (The Truth)
        selected_columns = [col for col in all_columns if col not in st.session_state.hidden_columns]
    
        if selected_columns:
            data = data[selected_columns]
        else:
            st.warning("No columns selected!")
            st.stop()
    
        # Build Grid Options
        gb = GridOptionsBuilder.from_dataframe(data)
        # gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20) # Add pagination
        gb.configure_side_bar() # Add a sidebar
        gb.configure_default_column(
            groupable=True, 
            value=True, 
            enableRowGroup=True, 
            aggFunc='sum', 
            editable=False,
            filterable=True,
            sortable=True,
            filter=True, # Enable filtering on all columns
            minWidth=100, # Ensure columns are readable
            resizable=True,
            filterParams={
                "buttons": ["apply", "reset"],
                "closeOnApply": True
            }
        )
        
        gridOptions = gb.build()
    
        # Placeholder for row count
        count_placeholder = st.empty()
    
        grid_response = AgGrid(
            data, 
            gridOptions=gridOptions, 
            enable_enterprise_modules=True, # Needed for some advanced filtering, but basic works
            update_mode=GridUpdateMode.MODEL_CHANGED,
            data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
            fit_columns_on_grid_load=False,
            theme='streamlit', # Add theme
            height=800, # Make the grid taller for scrolling
            key="report_grid"
        )
        
        # Show row count in the placeholder
        filtered_count = len(grid_response['data'])
        count_placeholder.info(f"Showing **{filtered_count}** entries")
        
        # --- Persist Filtered Data for MTC Tab ---
        # This allows us to use the filtered Heat List in the MTC dropdowns
        if len(grid_response['data']) > 0:
             st.session_state.filtered_df = pd.DataFrame(grid_response['data'])
        else:
             st.session_state.filtered_df = pd.DataFrame()

def render_mtc_viewer():
    st.header("MTC Report")
    
    MTC_FILE_PATH = r"E:\M537-02A26 - 1.xlsx"
    
    if os.path.exists(MTC_FILE_PATH):
        try:
            # Load Data without header clearly
            df = pd.read_excel(MTC_FILE_PATH, header=None)
            
            # --- Parsing Logic (Hardcoded for this specific template) ---
            # Row 1: Customer
            # Row 7: Part Details "Part Name: ... ; Part No:... ; Date code:..."
            # Row 13+: Data
            
            st.markdown("### 1. Edit Report Details")
            
            col1, col2 = st.columns(2)
            with col1:
                # Extract defaults safely
                try:
                    raw_part_line = str(df.iloc[7, 1]) if len(df) > 7 else "" # "Part Name: TB... ; Part No:..."
                except: raw_part_line = "Part Name: ; Part No: ; Date code:"
                
                part_details = st.text_input("Part Details Line", value=raw_part_line)
                customer_name = st.text_input("Customer Name", value="SUNDRAM FASTENERS LTD") # Placeholder/Extracted
                report_no_ref = st.text_input("Reference / Report Type", value="REFERENCE - Ductile iron J434C GRADE 4512")
                
            with col2:
                invoice_no = st.text_input("Invoice No", value="")
                date_str = st.text_input("Dispatch Date", value="")
                qty_str = st.text_input("Dispatch Quantity", value="no's")

            st.markdown("### 2. Edit Test Results")
            
            # Extract Chemistry Table safely
            # Assuming standard structure: Element | Spec | Heat1 | Heat2 ...
            # We will create a fresh dataframe for editing to be safe
            
            # Identifying Heat Numbers from Row 12 (index 12)
            try:
                # Potential Candidate 1 at Index 2
                candidate1 = str(df.iloc[12, 2]) if len(df) > 12 and pd.notna(df.iloc[12,2]) else "Heat 1"
                
                # Check if this candidate is "Percentage" (which is the Spec column header, not Heat No)
                if "percent" in candidate1.lower():
                    # Shift reading to right (Index 3 and 4)
                    parsed_heat1 = str(df.iloc[12, 3]) if len(df) > 12 and pd.notna(df.iloc[12,3]) else "Heat 1"
                    parsed_heat2 = str(df.iloc[12, 4]) if len(df) > 12 and pd.notna(df.iloc[12,4]) else "Heat 2"
                else:
                    parsed_heat1 = candidate1
                    parsed_heat2 = str(df.iloc[12, 3]) if len(df) > 12 and pd.notna(df.iloc[12,3]) else "Heat 2"
            except:
                parsed_heat1, parsed_heat2 = "42A26-2T", "43A26-2T"

            # --- Heat Number Dropdowns ---
            # Populate from Filtered Data in 'Report Viewer'
            available_heats = []
            available_samples = []
            
            fdf = st.session_state.filtered_df if 'filtered_df' in st.session_state and not st.session_state.filtered_df.empty else pd.DataFrame()
            
            if not fdf.empty:
                # Find Column Names
                heat_col = next((col for col in fdf.columns if "heat" in col.lower() and "no" in col.lower()), None)
                sample_col = next((col for col in fdf.columns if "sample" in col.lower() and "id" in col.lower()), None)
                
                # Get all unique heats initially
                if heat_col:
                    available_heats = fdf[heat_col].astype(str).unique().tolist()
                
                # Sample Id Filter
                if sample_col:
                     st.write("---") # Visual separator
                     available_samples = fdf[sample_col].astype(str).unique().tolist()
                     selected_sample = st.selectbox("Filter by Sample Id (Optional)", ["All"] + available_samples)
                     
                     if selected_sample != "All" and heat_col:
                         # Filter available heats based on selected sample
                         subset = fdf[fdf[sample_col].astype(str) == selected_sample]
                         if not subset.empty:
                             available_heats = subset[heat_col].astype(str).unique().tolist()
                             st.success(f"Sample **{selected_sample}** belongs to Heat: **{', '.join(available_heats)}**")

            # Prepare options
            # If parsed_heat matches one of available, good. If not, and we filtered, maybe default to the first available?
            # Let's keep parsed as an option but if available_heats has changed, the user sees mostly those.
            
            options1 = [parsed_heat1] + [h for h in available_heats if h != parsed_heat1]
            options2 = [parsed_heat2] + [h for h in available_heats if h != parsed_heat2]
            
            st.markdown("#### Select Heat Numbers")
            cols_h = st.columns(2)
            with cols_h[0]:
                heat1 = st.selectbox("Heat No 1 (Column 1)", options=options1, key="h1_sel")
            with cols_h[1]:
                heat2 = st.selectbox("Heat No 2 (Column 2)", options=options2, key="h2_sel")

            # --- Dynamic Data Parsing ---
            # Mapping Element Name -> Data Column
            element_map = {
                "Carbon": "C [%]", 
                "Silicon": "Si [%]", 
                "Manganese": "Mn [%]", 
                "Phosphorus": "P [%]", 
                "Sulphur": "S [%]", 
                "Copper": "Cu [%]", 
                "Nickel": "Ni [%]", 
                "Chromium": "Cr [%]", 
                "Moly": "Mo [%]", 
                "Magnesium": "Mg",
                "CE": "CE%",
                "Tin": "Sn%"
            }
            
            # Helper to get value safely
            def get_val(df_source, heat_val, index_offset=0):
                if df_source is not None and not df_source.empty and heat_val:
                    # Filter by Heat No (Find exact column name case-insensitive)
                    heat_col = next((c for c in df_source.columns if "heat" in c.lower() and "no" in c.lower()), None)
                    
                    if heat_col:
                        matches = df_source[df_source[heat_col].astype(str) == str(heat_val)]
                        if len(matches) > index_offset:
                            return matches.iloc[index_offset]
                return None

            # Get source data
            fdf = st.session_state.filtered_df if 'filtered_df' in st.session_state else None
            
            # Fetch Rows
            # For data_row_1: Always taken index 0 (first match)
            data_row_1 = get_val(fdf, heat1, 0)
            
            # For data_row_2: 
            # If Heat1 == Heat2, take index 1 (second match) to show the second sample.
            # Else, take index 0 (first match of that different heat).
            offset_2 = 1 if (heat1 == heat2) else 0
            data_row_2 = get_val(fdf, heat2, offset_2)

            # Build Table Data
            chem_data = []
            
            # Define Specs (Default values, editable later)
            specs = {
                "Carbon": "3.20 ~ 4.10%", "Silicon": "1.80 ~ 3.00%", "Manganese": "0.1 ~ 1.00%",
                "Phosphorus": "0.050% Max", "Sulphur": "0.035% Max", "Copper": "-", 
                "Nickel": "-", "Chromium": "-", "Moly": "-", "Magnesium": "0.025 ~ 0.060%",
                "CE": "-", "Tin": "-"
            }
            
            for elem, col_name in element_map.items():
                spec_val = specs.get(elem, "-")
                
                # Helper to format value
                def fmt_val(row, col):
                    if row is not None and col in row:
                        raw = row[col]
                        if pd.isna(raw): return ""
                        # Format floats to 3 decimal places? Or keep as is? User showed 3.49% (2 decimals)
                        # Let's try to keep 2-3 decimals if float.
                        try:
                            val_str = f"{float(raw):.2f}%" 
                        except:
                            val_str = str(raw)
                        
                        # Add % if missing and looks like a number
                        if "%" not in val_str and val_str.replace('.','',1).isdigit(): val_str += "%"
                        return val_str
                    return ""

                chem_data.append({
                    "Element": elem,
                    "Spec": spec_val,
                    "heat1_val": fmt_val(data_row_1, col_name),
                    "heat2_val": fmt_val(data_row_2, col_name),
                    "Hide": False
                })
            
            chem_df = pd.DataFrame(chem_data)
            
            edited_chem_df = st.data_editor(
                chem_df, 
                use_container_width=True, 
                num_rows="dynamic",
                column_config={
                    "heat1_val": st.column_config.TextColumn(label=heat1),
                    "heat2_val": st.column_config.TextColumn(label=heat2),
                    "Hide": st.column_config.CheckboxColumn("Hide", width="small")
                }
            )
            
            # Mechanical Properties
            st.markdown("#### Mechanical Properties")
            mech_data = [
                {"Parameter": "3.1 Hardness", "Spec": "156-217 HB", "heat1_val": "197/197/197/207/207 BHN", "heat2_val": "", "Hide": False},
                {"Parameter": "3.2 Tensile Strength", "Spec": "Min 450 Mpa", "heat1_val": "515.28 Mpa", "heat2_val": "", "Hide": False},
                {"Parameter": "3.3 Yield Strength", "Spec": "Min 295 Mpa", "heat1_val": "326.02 Mpa", "heat2_val": "", "Hide": False},
                {"Parameter": "3.4 % Of Elongation", "Spec": "Min 12 %", "heat1_val": "14.00%", "heat2_val": "", "Hide": False},
            ]
            mech_df = pd.DataFrame(mech_data)
            edited_mech_df = st.data_editor(
                mech_df, 
                use_container_width=True, 
                num_rows="dynamic",
                column_config={
                    "heat1_val": st.column_config.TextColumn(label=heat1),
                    "heat2_val": st.column_config.TextColumn(label=heat2),
                    "Hide": st.column_config.CheckboxColumn("Hide", width="small")
                }
            )
            
            
            # --- HTML Generation ---
            
            # Construct rows for chemistry
            chem_rows = ""
            for index, row in edited_chem_df.iterrows():
                if row.get("Hide", False): continue
                
                # Safe access using internal keys
                val1 = row["heat1_val"] if "heat1_val" in row and pd.notna(row["heat1_val"]) else ""
                val2 = row["heat2_val"] if "heat2_val" in row and pd.notna(row["heat2_val"]) else ""
                
                chem_rows += f"""
                <tr>
                    <td>{row['Element']}</td>
                    <td style='text-align: center;'>{row['Spec']}</td>
                    <td style='text-align: center; font-weight: bold;'>{val1}</td>
                    <td style='text-align: center; font-weight: bold;'>{val2}</td>
                </tr>
                """

            # Construct rows for mechanical
            mech_rows = ""
            for index, row in edited_mech_df.iterrows():
                 if row.get("Hide", False): continue

                 val1 = row["heat1_val"] if "heat1_val" in row and pd.notna(row["heat1_val"]) else ""
                 # val2 = row["heat2_val"] if "heat2_val" in row and pd.notna(row["heat2_val"]) else ""
                 
                 mech_rows += f"""
                <tr>
                    <td style='font-weight: bold;'>{row['Parameter']}</td>
                    <td style='text-align: center;'>{row['Spec']}</td>
                    <td style='text-align: center;' colspan='2'>{val1}</td> 
                </tr>
                """

            s = st.session_state.app_settings
            h_color = s.get("header_fill_color", "#d9e1f2")
            b_style = s.get("border_style", "thin")
            f_family = s.get("font_family", "Calibri")
            f_size = s.get("font_size", 10)
            h_align = s.get("header_align", "center")
            border_px = "1px" if b_style == "thin" else "2px" if b_style == "medium" else "3px" if b_style == "thick" else "0"

            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
            <title>MTC Report</title>
            <style>
                @page {{ size: A4; margin: 10mm; }}
                
                /* Dynamic Styles from Settings */

                body {{ 
                    font-family: "{f_family}", "Arial", sans-serif; 
                    background-color: #f0f0f0; 
                    margin: 0; 
                    padding: 20px; 
                }}
                .page {{
                    width: 210mm;
                    min-height: 297mm;
                    padding: 10mm;
                    margin: 0 auto;
                    background: white;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                    border: 1px solid #ccc;
                    box-sizing: border-box;
                }}
                table {{ width: 100%; border-collapse: collapse; font-size: {f_size}pt; }}
                td, th {{ border: {border_px} solid black; padding: 4px; vertical-align: middle; }}
                
                .logo-cell {{ 
                    background-color: #000080; 
                    color: white; 
                    text-align: center; 
                    width: 15%; 
                    vertical-align: middle;
                    padding: 5px;
                }}
                
                /* Hexagon Logo Construction */
                .logo-container {{
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    height: 100%;
                }}
                
                .tvs-hex {{
                    width: 80px;
                    height: 55px;
                    background-color: white;
                    position: relative;
                    /* Create Hexagon shape */
                    clip-path: polygon(20% 0%, 80% 0%, 100% 50%, 80% 100%, 20% 100%, 0% 50%);
                    display: flex;
                    justify-content: center;
                    align-items: center;
                }}
                
                .tvs-oval {{
                    width: 74px;
                    height: 48px;
                    background-color: #000080;
                    border-radius: 50%; /* Oval */
                    display: flex;
                    justify-content: center;
                    align-items: center;
                }}
                
                .tvs-text {{
                    color: white;
                    font-weight: bold;
                    font-family: "Arial Narrow", Arial, sans-serif;
                    font-size: 24pt;
                    letter-spacing: 1px;
                    /* Slight skew/italic if needed, but plain looks close */
                }}
                
                .company-header {{ text-align: center; color: #000000; }}
                .company-name {{ font-weight: bold; font-size: 14pt; color: #0070c0; }}
                .lab-report-title {{ font-weight: bold; margin-top: 5px; font-size: 12pt; }}
                
                .info-cell {{ width: 35%; font-size: 9pt; }}
                .info-line {{ margin-bottom: 2px; }}
                
                .section-header {{ 
                    background-color: {h_color}; 
                    font-weight: bold; 
                    text-align: {h_align}; 
                    padding: 4px; 
                }}
                
                .obs-header {{ background-color: white; }}
                .chem-header {{ background-color: #f2f2f2; font-weight: bold; text-align: center; }}
                
                .footer-table td {{ height: 50px; vertical-align: bottom; text-align: center; }}
                .footer-label {{ background-color: #f2f2f2; height: 20px; vertical-align: middle; font-weight: bold; }}
                
                @media print {{
                    body {{ background: none; padding: 0; margin: 0; }}
                    .page {{ 
                        box-shadow: none; 
                        margin: 0; 
                        border: none; 
                        width: 100%; 
                        height: auto; 
                        padding: 0; 
                    }}
                    .tvs-hex {{ -webkit-print-color-adjust: exact; }}
                    .tvs-oval {{ -webkit-print-color-adjust: exact; }}
                    .logo-cell {{ -webkit-print-color-adjust: exact; }}
                    .no-print {{ display: none; }}
                    html, body {{ height: 100%; }}
                }}
            </style>
            </head>
            <body>
            
            <div class="page">
                <!-- Header -->
                <table>
                    <tr>
                        <td class="logo-cell">
                             <div class="logo-container">
                                 <div class="tvs-hex">
                                     <div class="tvs-oval">
                                         <span class="tvs-text">TVS</span>
                                     </div>
                                 </div>
                             </div>
                        </td>
                        <td class="company-header">
                            <div class="company-name">AUTOLEC DIVISION-FOUNDRY</div>
                            <div>Gummidipoondi-601201</div>
                            <div>E-Mail: mythili.g@sfl.co.in</div>
                            <div class="lab-report-title">LABORATORY REPORT</div>
                        </td>
                        <td class="info-cell">
                            <div class="info-line"><b>Customer Part No:</b> A 400 351 02 11</div>
                            <div class="info-line"><b>Invoice No :</b> {invoice_no}</div>
                            <div class="info-line"><b>Despatch Quantity:</b> {qty_str}</div>
                            <div class="info-line"><b>Dispatch Date :</b> {date_str}</div>
                        </td>
                    </tr>
                </table>
                
                <!-- Reference & Part -->
                <div style="border: 1px solid black; border-top: none; padding: 4px; font-weight: bold; font-size: 10pt;">
                    {report_no_ref}
                </div>
                <div style="border: 1px solid black; border-top: none; padding: 4px; font-weight: bold; font-size: 10pt; margin-bottom: 10px;">
                    {part_details}
                </div>
                
                <!-- Main Data Table -->
                <table>
                    <tr>
                        <th style="width: 30%;">PARAMETER</th>
                        <th style="width: 30%;">SPECIFICATION</th>
                        <th colspan="2">OBSERVATIONS</th>
                    </tr>
                    
                    <!-- Chemical Section -->
                    <tr>
                        <td colspan="4" class="section-header">Specification <br> {st.session_state.app_settings.get('chem_title', '1. Chemical composition')}</td>
                    </tr>
                    
                    <tr>
                        <td colspan="2" style="border: none;"></td> <!-- Spacer/Layout check -->
                        <td colspan="2" style="text-align: center; font-weight: bold; border-left: 1px solid black; border-bottom: 0;">Heat No:</td>
                    </tr>
                    
                    <tr class="chem-header">
                        <td>Element</td>
                        <td>Percentage</td>
                        <td>{heat1}</td>
                        <td>{heat2}</td>
                    </tr>
                    
                    {chem_rows}
                    
                    <!-- Mechanical Section -->
                    <tr>
                        <td colspan="4" class="section-header">{st.session_state.app_settings.get('mech_title', '2. Mechanical Properties')}</td>
                    </tr>
                    
                    {mech_rows}
                    
                    <!-- Microstructure Section -->
                     <tr>
                        <td colspan="4" class="section-header">{st.session_state.app_settings.get('micro_title', '3. Microstructure')}</td>
                    </tr>
                    <tr>
                        <td colspan="2" style="padding: 10px; text-align: center;">Thar Graphite from shall be 80 percent Types I & II as determined in accordance with ASTM A 247</td>
                        <td colspan="2" style="padding: 10px; text-align: center;">Graphite form Type V and VI,<br>Nodularity: 90%<br>Nodule count: 290/mm²</td>
                    </tr>
                    
                     <!-- Matrix Section -->
                     <tr>
                        <td colspan="4" class="section-header">{st.session_state.app_settings.get('matrix_title', '3.1 Matrix')}</td>
                    </tr>
                    <tr>
                        <td colspan="2" style="padding: 20px; text-align: center;">Ferrite - Pearlite</td>
                        <td colspan="2" style="padding: 20px; text-align: center;">Predominantly Ferrite matrix with Pearlite</td>
                    </tr>
                    
                     <!-- Images -->
                    <tr>
                         <td colspan="2" style="height: 120px; text-align: center; vertical-align: bottom; font-weight: bold;">
                            <!-- Placeholder for image -->
                            <div style="background: #eee; height: 100px; width: 100px; margin: auto;">Image</div>
                            Polished Image at 100X
                         </td>
                         <td colspan="2" style="height: 120px; text-align: center; vertical-align: bottom; font-weight: bold;">
                             <div style="background: #eee; height: 100px; width: 100px; margin: auto;">Image</div>
                             Etched Image at 100X
                         </td>
                    </tr>

                </table>
                
                <div style="padding: 5px; border: 1px solid black; border-top: none; font-weight: bold; font-size: 10pt;">
                    Conclusion: The above material is satisfactory to Ductile iron J434C GRADE 4512.
                </div>
                
                <div style="margin-top: 10px;">
                     <table class="footer-table">
                        <tr>
                            <td rowspan="2" style="width: 25%; text-align: left; vertical-align: middle; padding-left: 10px;">SFL/FD/9.15</td>
                            <td class="footer-label">REPORTED BY</td>
                            <td class="footer-label">APPROVED BY</td>
                        </tr>
                        <tr>
                            <td style="height: 60px;">G.Mythili</td>
                            <td style="height: 60px;">T.Thirugnanam</td>
                        </tr>
                     </table>
                 </div>
                 
            </div>
            
            </body>
            </html>
            """
            
            st.markdown("### Preview & Print")
            
            # Action Buttons
            col_print, col_download = st.columns(2)
            
            with col_print:
                if st.button("Generate Print View"):
                    st.components.v1.html(html_content, height=1000, scrolling=True)
                    
                    with open("report_print.html", "w", encoding="utf-8") as f:
                        f.write(html_content + "<script>window.print()</script>")
                    
                    st.success("Print View Generated.")
                    # We can't automatically open tabs safely in all deployments, but valid in local
                    st.info("Click 'Open in New Tab' to print, or 'Download' to save.")
            
            with col_download:
                # Helper function to generate Excel
                def generate_excel():
                    output = io.BytesIO()
                    try:
                        from openpyxl import load_workbook
                        from openpyxl.cell.cell import Cell, MergedCell
                        from openpyxl.styles import PatternFill, Border, Side, Alignment

                        # 1. Load Initial from Settings
                        tpl_path = st.session_state.app_settings.get("mtc_template_path", DEFAULT_MTC_TEMPLATE)
                        wb = load_workbook(tpl_path)
                        ws = wb.active

                        # 2. Get Settings
                        s = st.session_state.app_settings
                        h_align = s.get("header_align", "center")

                        # --- Pure Template Mode: Preserve Layout ---
                        # ws.page_setup.paperSize = ws.PAPERSIZE_A4
                        # ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                        # ws.page_setup.fitToWidth = 1
                        # ws.page_setup.fitToHeight = 0
                        # ws.sheet_properties.pageSetUpPr.fitToPage = True
                        
                        # ws.page_margins.left = 0.3
                        # ws.page_margins.right = 0.3
                        # ws.page_margins.top = 0.5
                        # ws.page_margins.bottom = 0.5

                        # ws.column_dimensions['A'].width = 4
                        # ws.column_dimensions['B'].width = 28
                        # ws.column_dimensions['C'].width = 18
                        # ws.column_dimensions['D'].width = 16
                        # ws.column_dimensions['E'].width = 16

                        # 2. Proactively Unmerge Data Area (Removed to preserve alignment)
                        # We will handle merges dynamically in write_styled
                        pass

                        # --- Pure Template Mode: No Row Shifts ---
                        extra_chem_rows = 0 
                        
                        # CLEAN SLATE: Just clear values in data area safely
                        try:
                            # Use a super-safe approach: try-except every cell to avoid MergedCell locks
                            for r in range(15, 61):
                                for c in range(1, 6):
                                    try:
                                        target = ws.cell(row=r, column=c)
                                        # If it's a MergedCell, setting .value will raise AttributeError or similar
                                        target.value = None
                                    except:
                                        pass
                        except: pass

                        # --- Disable Unstable Feature Clearing ---
                        # try:
                        #     ws.conditional_formatting = type(ws.conditional_formatting)()
                        # except: pass
                            
                        # --- Clear Tables (Remove Orange Banding) ---            

                        # try:
                        #     if hasattr(ws, 'tables'):
                        #         for table_name in list(ws.tables.keys()):
                        #             del ws.tables[table_name]
                        # except: pass

                        # --- Define Styles from Settings ---
                        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
                        
                        b_style = s.get("border_style", "thin")
                        if b_style == "none":
                            full_border = Border()
                        else:
                            side_obj = Side(border_style=b_style, color="000000")
                            full_border = Border(top=side_obj, left=side_obj, right=side_obj, bottom=side_obj)
                            
                        custom_font = Font(name=s.get("font_family", "Calibri"), size=s.get("font_size", 10))
                        
                        # Process HEX color for openpyxl
                        h_hex = s.get("header_fill_color", "#d9e1f2").replace("#", "")
                        header_fill = PatternFill(start_color=h_hex, end_color=h_hex, fill_type="solid")
                        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        
                        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

                        def write_styled(row, col, value, align='center', fill=False, is_header=False):
                            try:
                                cell = ws.cell(row=row, column=col)
                                target = cell
                                
                                # Handle Merged Cells: Resolve to Master
                                for rng in ws.merged_cells.ranges:
                                    if cell.coordinate in rng:
                                        target = ws.cell(row=rng.min_row, column=rng.min_col)
                                        break
                                
                                # Final Safety: Try to write. If it fails due to MergedCell, we skip or pass.
                                try:
                                    target.value = value
                                except:
                                    # If it's still a MergedCell (read-only), we can't write to it.
                                    # In a 'Pure Filling' mode, we just pass to avoid crashing.
                                    pass

                                # Styles
                                target.border = full_border
                                target.font = custom_font
                                if align == 'center': target.alignment = center_align
                                elif align == 'left': target.alignment = left_align
                                elif align == 'right': target.alignment = right_align
                                
                                if is_header: target.fill = header_fill
                                elif fill: target.fill = white_fill
                                else: target.fill = PatternFill(fill_type=None)
                            except: pass

                        def safe_write(row, col, value):
                            try:
                                cell = ws.cell(row=row, column=col)
                                target = cell
                                for rng in ws.merged_cells.ranges:
                                    if cell.coordinate in rng:
                                        target = ws.cell(row=rng.min_row, column=rng.min_col)
                                        break
                                try:
                                    target.value = value
                                except: pass
                            except: pass

                        # --- Update Headers ---
                        # Use safe_write to handle potential merges in template headers
                        if invoice_no: safe_write(4, 5, invoice_no)
                        if qty_str:    safe_write(5, 5, qty_str)
                        if date_str:   safe_write(6, 5, date_str)
                        if part_details: safe_write(8, 2, part_details)
                        
                        # Write Heat Numbers to Header Row (Row 13)
                        write_styled(13, 4, heat1 if heat1 else "", fill=True)
                        write_styled(13, 5, heat2 if heat2 else "", fill=True)
                        
                        # --- Insert Logo (Only if template is empty) ---
                        try:
                            if len(ws._images) == 0:
                                from openpyxl.drawing.image import Image
                                # Using the specific uploaded logo path
                                logo_path = r"C:/Users/Admin/.gemini/antigravity/brain/8e5bd36d-b8ec-4d97-baca-df8e1ef9f9df/uploaded_image_1_1768116930343.png"
                                if os.path.exists(logo_path):
                                    img = Image(logo_path)
                                    img.height = 120
                                    img.width = 75 
                                    ws.add_image(img, 'A1')
                        except Exception as logo_err:
                            print(f"Could not add logo: {logo_err}")

                        # --- Section 1: Chemistry ---
                        current_row = 14
                        header_text = f"Specification\n{s.get('chem_title', '1. Chemical composition')}"
                        write_styled(current_row, 2, header_text, align=h_align, is_header=True)
                        try:
                            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
                        except: pass
                        current_row = 15 # Data starts at 15
                             
                        # Write Chemistry Data
                        for _, row in edited_chem_df.iterrows():
                             if row.get("Hide", False): continue
                             
                             # Col 1: Clear
                             write_styled(current_row, 1, "", align='center')
                             # Col 2: Element
                             write_styled(current_row, 2, row['Element'], align='left')
                             # Col 3: Spec
                             write_styled(current_row, 3, row['Spec'])
                             # Col 4 & 5: Heat Values
                             write_styled(current_row, 4, row.get('heat1_val', ''), fill=True)
                             write_styled(current_row, 5, row.get('heat2_val', ''), fill=True)
                             current_row += 1
                                  
                        # --- Section 2: Mechanical Properties ---
                        # Add a tiny gap if needed, or just start next row
                        
                        # Header Row
                        try:
                            for c in range(1, 6):
                                cell_chk = ws.cell(row=current_row, column=c)
                                for rng in list(ws.merged_cells.ranges):
                                    if cell_chk.coordinate in rng:
                                        try: ws.unmerge_cells(str(rng))
                                        except: pass
                            # Clear A, Merge B-E
                            write_styled(current_row, 1, "") 
                            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
                        except: pass
                        write_styled(current_row, 2, s.get("mech_title", "2. Mechanical Properties"), align=h_align, is_header=True)
                        current_row += 1

                        # Mechanical Data
                        for _, row in edited_mech_df.iterrows():
                            if row.get("Hide", False): continue
                            
                            # Explicitly clear/unmerge data row to prevent Column A overlap
                            try:
                                for c in range(1, 6): 
                                    cell_chk = ws.cell(row=current_row, column=c)
                                    for rng in list(ws.merged_cells.ranges):
                                        if cell_chk.coordinate in rng:
                                            try: ws.unmerge_cells(str(rng))
                                            except: pass
                            except: pass

                            # Col 1: Clear
                            write_styled(current_row, 1, "", align='left')
                            # Col 2: Parameter
                            write_styled(current_row, 2, row['Parameter'], align='left')
                            # Col 3: Spec
                            write_styled(current_row, 3, row['Spec'])
                            # Col 4: Heat 1 Value
                            write_styled(current_row, 4, row.get('heat1_val', ''), fill=True)
                            # Col 5: Clear/Border
                            write_styled(current_row, 5, "") 
                            
                            current_row += 1
                        
                        # --- Section 3: Microstructure ---
                        write_styled(current_row, 2, s.get("micro_title", "3. Microstructure"), align=h_align, is_header=True)
                        try: ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
                        except: pass
                        current_row += 1
                        
                        micro_val1 = "Thar Graphite from shall be 80 percent Types I & II as determined in accordance with ASTM A 247"
                        micro_val2 = "Graphite form Type V and VI,\nNodularity: 90%\nNodule count: 290/mm²"
                        write_styled(current_row, 1, "") # Col 1 Clear
                        write_styled(current_row, 2, micro_val1, align='center')
                        try: ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
                        except: pass
                        write_styled(current_row, 4, micro_val2, align='center')
                        try: ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
                        except: pass
                        ws.row_dimensions[current_row].height = 60
                        current_row += 1

                        # --- Section 3.1: Matrix ---
                        write_styled(current_row, 2, s.get("matrix_title", "3.1 Matrix"), align=h_align, is_header=True)
                        try: ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
                        except: pass
                        current_row += 1
                        
                        matrix_val1 = "Ferrite - Pearlite"
                        matrix_val2 = "Predominantly Ferrite matrix with Pearlite"
                        write_styled(current_row, 1, "") # Col 1 Clear
                        write_styled(current_row, 2, matrix_val1, align='center')
                        try: ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
                        except: pass
                        write_styled(current_row, 4, matrix_val2, align='center')
                        try: ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
                        except: pass
                        ws.row_dimensions[current_row].height = 40
                        current_row += 1

                        # --- Conclusion ---
                        current_row += 1
                        try: ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                        except: pass
                        
                        # Attempt to extract Grade from Part Details for dynamic conclusion
                        target_grade = "GRADE 4512" 
                        if part_details and "GRADE" in part_details.upper():
                            try: target_grade = part_details.split("GRADE")[-1].strip()
                            except: pass
                        
                        conclusion_text = f"Conclusion: The above material is satisfactory to Ductile iron J434C GRADE {target_grade}."
                        write_styled(current_row, 1, conclusion_text, align='left')
                        ws.row_dimensions[current_row].height = 30
                        current_row += 1
                        
                        # --- Footer Table ---
                        # Standardize Footer Row Height
                        ws.row_dimensions[current_row].height = 25
                        ws.row_dimensions[current_row+1].height = 45 

                        # Labels Row
                        write_styled(current_row, 1, "SFL/FD/9.15", align='left')
                        
                        # Reported By (Col 2-3)
                        write_styled(current_row, 2, "REPORTED BY", is_header=True)
                        try: ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
                        except: pass
                        # Ensure cell 3 also has border for merge consistency
                        ws.cell(row=current_row, column=3).border = full_border
                        
                        # Approved By (Col 4-5)
                        write_styled(current_row, 4, "APPROVED BY", is_header=True)
                        try: ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
                        except: pass
                        ws.cell(row=current_row, column=5).border = full_border
                        
                        current_row += 1

                        # Signature/Names Row
                        write_styled(current_row, 1, "", align='center') # Bordered empty cell for code 
                        
                        # Mythili (Col 2-3)
                        write_styled(current_row, 2, "G.Mythili", align='center')
                        try: ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
                        except: pass
                        ws.cell(row=current_row, column=3).border = full_border

                        # Thirugnanam (Col 4-5)
                        write_styled(current_row, 4, "T.Thirugnanam", align='center')
                        try: ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
                        except: pass
                        ws.cell(row=current_row, column=5).border = full_border
                                
                        wb.save(output)
                        return output.getvalue()
                    except Exception as e:
                        st.error(f"Error generating Excel: {e}")
                        return None

                excel_data = generate_excel()
                
                if excel_data:
                     st.download_button(
                        label="Download Excel (.xlsx)",
                        data=excel_data,
                        file_name="MTC_Report_Modified.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Download the report as an editable Excel file."
                    )
                     
            
            # Persistent link helper
            if os.path.exists("report_print.html"):
                 if st.button("Open in New Tab for Printing"):   
                        webbrowser.open("file://" + os.path.realpath("report_print.html"))
                
        except Exception as e:
            st.error(f"Error loading MTC file: {e}")
    else:
        st.error(f"MTC File not found at: {MTC_FILE_PATH}")

def main():
    st.title("Spectro Report Viewer")
    
    tab1, tab2 = st.tabs(["Report Viewer", "MTC"])
    
    with tab1:
        render_report_viewer()
    
    with tab2:
        render_mtc_viewer()

if __name__ == "__main__":
    main()
