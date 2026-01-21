from openpyxl import load_workbook
import os

template_path = r"e:\2026 Report Famat\Final correct.xlsx"
if not os.path.exists(template_path):
    print("Template not found!")
else:
    wb = load_workbook(template_path)
    ws = wb.active
    print(f"Sheet Name: {ws.title}")
    print("\nMerged Ranges:")
    for rng in sorted(list(ws.merged_cells.ranges), key=lambda x: str(x)):
        print(f" - {rng}")
    
    print("\nColumn Widths:")
    for col in ['A', 'B', 'C', 'D', 'E']:
        width = ws.column_dimensions[col].width
        print(f" - Col {col}: {width}")

    print("\nSample Values in Specific Rows:")
    checks = [(4,3), (5,3), (6,3), (8,1), (13,1), (13,2), (13,3), (13,4)]
    for r, c in checks:
        val = ws.cell(row=r, column=c).value
        print(f" - Row {r}, Col {c}: {val}")
